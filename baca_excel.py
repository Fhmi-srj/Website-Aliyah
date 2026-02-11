"""
Script untuk membaca seluruh isi dari:
  - ABSEN SISWA REAL.xlsx
  - MENGAJAR REAL.xlsx
Menggunakan openpyxl untuk membaca setiap sheet, setiap baris, setiap kolom.
"""

import openpyxl
import os
import sys

def read_excel_full(filepath):
    """Membaca seluruh isi file Excel dengan detail lengkap."""
    if not os.path.exists(filepath):
        print(f"❌ File tidak ditemukan: {filepath}")
        return

    print("=" * 120)
    print(f"📄 FILE: {os.path.basename(filepath)}")
    print(f"   Path: {filepath}")
    print(f"   Size: {os.path.getsize(filepath):,} bytes")
    print("=" * 120)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"   Jumlah Sheet: {len(wb.sheetnames)}")
    print(f"   Nama Sheet: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("-" * 120)
        print(f"📋 SHEET: '{sheet_name}'")
        print(f"   Dimensi: {ws.dimensions}")
        print(f"   Baris: {ws.max_row}, Kolom: {ws.max_column}")
        print(f"   Merged Cells: {list(ws.merged_cells.ranges) if ws.merged_cells.ranges else 'Tidak ada'}")
        print("-" * 120)

        if ws.max_row is None or ws.max_row == 0:
            print("   (Sheet kosong)")
            print()
            continue

        # Print header info (kolom)
        print(f"\n   {'ROW':<6}", end="")
        for col in range(1, (ws.max_column or 0) + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"  | {col_letter:<20}", end="")
        print()
        print("   " + "-" * (6 + 24 * (ws.max_column or 1)))

        # Print setiap baris
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            row_num = row[0].row
            print(f"   {row_num:<6}", end="")
            for cell in row:
                value = cell.value
                if value is None:
                    display = ""
                elif isinstance(value, str):
                    display = value.strip()[:20]
                else:
                    display = str(value)[:20]
                print(f"  | {display:<20}", end="")
            print()

        print(f"\n   Total baris data: {ws.max_row}")
        print(f"   Total kolom: {ws.max_column}")
        print()

    wb.close()
    print()


def read_excel_raw(filepath):
    """Membaca seluruh isi file Excel dan print SEMUA data tanpa truncate."""
    if not os.path.exists(filepath):
        print(f"❌ File tidak ditemukan: {filepath}")
        return

    print("=" * 120)
    print(f"📄 FILE (RAW DETAIL): {os.path.basename(filepath)}")
    print("=" * 120)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"📋 SHEET: '{sheet_name}' | Baris: {ws.max_row} | Kolom: {ws.max_column}")
        print(f"{'='*80}")

        if ws.max_row is None or ws.max_row == 0:
            print("   (Sheet kosong)")
            continue

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
            row_num = row[0].row
            print(f"\n--- Baris {row_num} ---")
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                value = cell.value
                cell_type = type(value).__name__ if value is not None else "NoneType"
                print(f"   [{col_letter}{row_num}] ({cell_type}) = {repr(value)}")

    wb.close()
    print()


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))

    files = [
        os.path.join(base_dir, "ABSEN SISWA REAL.xlsx"),
        os.path.join(base_dir, "MENGAJAR REAL.xlsx"),
    ]

    # Mode: 'table' untuk tabel ringkas, 'raw' untuk detail penuh per cell
    mode = sys.argv[1] if len(sys.argv) > 1 else "raw"

    for f in files:
        if mode == "raw":
            read_excel_raw(f)
        else:
            read_excel_full(f)
